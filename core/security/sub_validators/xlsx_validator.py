"""
xlsx_validator — Validierung von XLSX / XLSM-Dateien.

XLSX ist ein ZIP-Container mit XML-Parts (Office Open XML). Die Angriffs-
fläche ist entsprechend groß:

Erkannte Angriffe:
  - **Zip-Bomb**: Dekomprimierte Größe oder Eintragsanzahl außerhalb
    sinnvoller Grenzen. → HIGH bzw. CRITICAL.
  - **Formula-Injection (CWE-1236)**: ``=cmd|'/c calc'!A1``,
    ``=WEBSERVICE``, ``=IMPORTDATA`` u. a. → HIGH.
  - **VBA-Makros** (via ``oletools.olevba``): Jedes Makro in einer
    Klienten-Datei ist potenziell gefährlich. → HIGH.
  - **DDE-Links** (via ``oletools.msodde``): Legacy-Feature, das
    Kommandoausführung ermöglicht. → CRITICAL.
  - **Externe Links** (``xl/externalLinks/*``): PDF/HTTP-Verbindungen
    aus einer Tabelle sind in 99 % der Fälle ein Lieferanten-Phishing-
    Muster. → HIGH.
  - **XXE über XML-Parts**: defusedxml erzwingt sichere Parser.

Author: Patrick Riederich
Version: 1.0
"""

from __future__ import annotations

import logging
import zipfile
from pathlib import Path

from core.logger import get_logger
from core.security.magika_adapter import identify
from core.security.sub_validators.base import SubValidator
from core.security.validation_report import Severity, Threat, ValidationReport

_log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Konstanten
# ---------------------------------------------------------------------------

# Zip-Container-Grenzen (BSI-Empfehlung + Erfahrungswert)
MAX_ZIP_ENTRIES: int = 10_000
MAX_UNCOMPRESSED_BYTES: int = 500 * 1024 * 1024  # 500 MB
MAX_COMPRESSION_RATIO: float = 100.0  # kompr:unkompr — 1:100 ist realistisch
MIN_COMPRESSED_BYTES_FOR_RATIO_CHECK: int = 1024  # Kleine Dateien ignorieren

# Formula-Injection-Präfixe: Spreadsheet-Engines werten Zellen, die mit
# diesen Zeichen beginnen, als Formel aus (OWASP-Dokumentation).
FORMULA_PREFIXES: frozenset[str] = frozenset({"=", "+", "-", "@", "\t", "\r"})

# Besonders gefährliche Formel-Funktionen
DANGEROUS_FORMULAS: tuple[str, ...] = (
    "CMD",
    "DDE",
    "WEBSERVICE",
    "HYPERLINK",
    "IMPORTDATA",
    "IMPORTXML",
    "IMPORTRANGE",
    "IMPORTFEED",
    "IMPORTHTML",
    "CALL",
    "EXEC",
)

# Stichprobengrenze für Formula-Threats — eine Datei mit 10k Formeln soll
# nicht 10k Threats erzeugen.
MAX_FORMULA_SAMPLES: int = 10

# Pfad-Präfixe innerhalb des ZIP, die externe Verknüpfungen enthalten
EXTERNAL_LINK_PREFIX: str = "xl/externalLinks/"


class XlsxValidator(SubValidator):
    """Validierer für XLSX / XLSM — prüft Container, Makros, DDE, Formeln."""

    def validate(self, path: Path, report: ValidationReport) -> None:
        """Führt alle XLSX-spezifischen Prüfungen durch.

        Args:
            path: Zu prüfender Pfad.
            report: Report zum Anhängen.
        """
        # Magika: Muss wie ein OOXML/ZIP aussehen. "exe", "javascript" etc.
        # sind harte Spoofing-Befunde.
        ident = identify(path)
        if ident.label not in ("xlsx", "xlsm", "xlsb", "zip", "docx", "ooxml"):
            report.add(
                Threat(
                    code="XLSX_CONTENT_MISMATCH",
                    severity=Severity.CRITICAL,
                    message=(
                        f"Inhalt ist keine Excel-Datei — Magika erkannte "
                        f"'{ident.label}' ({ident.description})."
                    ),
                    context={
                        "detected_label": ident.label,
                        "mime_type": ident.mime_type,
                    },
                )
            )
            return  # Bei CRITICAL keine weiteren Checks

        # ZIP-Struktur öffnen
        try:
            zf = zipfile.ZipFile(path, "r")
        except (zipfile.BadZipFile, OSError) as exc:
            report.add(
                Threat(
                    code="XLSX_CORRUPT_CONTAINER",
                    severity=Severity.HIGH,
                    message=f"XLSX-Container nicht lesbar: {exc}",
                )
            )
            return

        with zf:
            self._check_zip_structure(zf, path, report)
            self._check_external_links(zf, report)

        # Makros (oletools) — nutzt den Datei-Pfad direkt, nicht ZipFile
        self._check_macros(path, report)
        # DDE-Links
        self._check_dde(path, report)
        # Formula-Injection-Scan
        self._check_formulas(path, report)

    # ------------------------------------------------------------------
    # Private Prüfer
    # ------------------------------------------------------------------

    @staticmethod
    def _check_zip_structure(
        zf: zipfile.ZipFile, path: Path, report: ValidationReport
    ) -> None:
        """Prüft Zip-Bomb-Indikatoren (Eintragsanzahl, Größe, Ratio)."""
        infos = zf.infolist()
        if len(infos) > MAX_ZIP_ENTRIES:
            report.add(
                Threat(
                    code="XLSX_ZIP_TOO_MANY_ENTRIES",
                    severity=Severity.CRITICAL,
                    message=(
                        f"ZIP enthält {len(infos)} Einträge "
                        f"(Limit: {MAX_ZIP_ENTRIES}) — Zip-Bomb-Indikator."
                    ),
                    context={"entries": len(infos), "limit": MAX_ZIP_ENTRIES},
                )
            )
            return

        total_uncompressed = sum(i.file_size for i in infos)
        if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
            report.add(
                Threat(
                    code="XLSX_ZIP_UNCOMPRESSED_TOO_LARGE",
                    severity=Severity.CRITICAL,
                    message=(
                        f"Dekomprimierte Größe "
                        f"{total_uncompressed // (1024 * 1024)} MB "
                        f"überschreitet Limit "
                        f"{MAX_UNCOMPRESSED_BYTES // (1024 * 1024)} MB — "
                        "Zip-Bomb-Indikator."
                    ),
                    context={
                        "uncompressed_bytes": total_uncompressed,
                        "limit_bytes": MAX_UNCOMPRESSED_BYTES,
                    },
                )
            )
            return

        total_compressed = path.stat().st_size
        if (
            total_compressed >= MIN_COMPRESSED_BYTES_FOR_RATIO_CHECK
            and total_compressed > 0
        ):
            ratio = total_uncompressed / total_compressed
            if ratio > MAX_COMPRESSION_RATIO:
                report.add(
                    Threat(
                        code="XLSX_ZIP_COMPRESSION_RATIO",
                        severity=Severity.HIGH,
                        message=(
                            f"Kompressionsverhältnis {ratio:.0f}:1 "
                            f"überschreitet Limit {MAX_COMPRESSION_RATIO:.0f}:1 — "
                            "Zip-Bomb-Verdacht."
                        ),
                        context={
                            "ratio": round(ratio, 1),
                            "limit": MAX_COMPRESSION_RATIO,
                        },
                    )
                )

    @staticmethod
    def _check_external_links(zf: zipfile.ZipFile, report: ValidationReport) -> None:
        """Meldet Einträge unter ``xl/externalLinks/`` als HIGH-Threat."""
        external = [n for n in zf.namelist() if n.startswith(EXTERNAL_LINK_PREFIX)]
        if external:
            report.add(
                Threat(
                    code="XLSX_EXTERNAL_LINKS",
                    severity=Severity.HIGH,
                    message=(
                        f"XLSX enthält {len(external)} externe Verknüpfung(en) — "
                        "typisches Phishing-Muster (Daten-Exfiltration via "
                        "remote-Tabelle)."
                    ),
                    context={"entries": external[:5]},
                )
            )

    @staticmethod
    def _check_macros(path: Path, report: ValidationReport) -> None:
        """Scannt VBA-Makros mittels ``oletools.olevba``."""
        # oletools schreibt lautstark auf den Root-Logger; dämpfen.
        vba_logger = logging.getLogger("olevba")
        orig_level = vba_logger.level
        vba_logger.setLevel(logging.ERROR)
        try:
            from oletools.olevba import VBA_Parser  # noqa: PLC0415

            parser = VBA_Parser(str(path))
            if parser.detect_vba_macros():
                macros = list(parser.extract_all_macros())
                report.add(
                    Threat(
                        code="XLSX_VBA_MACRO",
                        severity=Severity.HIGH,
                        message=(
                            f"VBA-Makro erkannt ({len(macros)} Modul(e)) — "
                            "Makros können Kommandoausführung bewirken."
                        ),
                        context={
                            "module_count": len(macros),
                        },
                    )
                )
            parser.close()
        except Exception as exc:  # noqa: BLE001 -- olevba/msodde/openpyxl koennen bei defekten Files unspezifizierte Errors werfen, Scanner darf nicht crashen
            # olevba kann bei defekten Dateien abbrechen — nur loggen
            _log.debug("olevba scan fehlgeschlagen für %s: %s", path.name, exc)
            report.add(
                Threat(
                    code="XLSX_MACRO_SCAN_ERROR",
                    severity=Severity.LOW,
                    message="Makro-Scan konnte nicht abgeschlossen werden.",
                    context={"error": str(exc)[:120]},
                )
            )
        finally:
            vba_logger.setLevel(orig_level)

    @staticmethod
    def _check_dde(path: Path, report: ValidationReport) -> None:
        """Scannt DDE-Links mittels ``oletools.msodde``."""
        try:
            from oletools import msodde  # noqa: PLC0415

            # msodde.process_file gibt einen String mit \n-separierten
            # DDE-Links zurück (leerer String = keine).
            dde = msodde.process_file(str(path), field_filter_mode="all")
            lines = [ln.strip() for ln in dde.splitlines() if ln.strip()]
            if lines:
                report.add(
                    Threat(
                        code="XLSX_DDE_LINK",
                        severity=Severity.CRITICAL,
                        message=(
                            f"DDE-Link erkannt ({len(lines)} Eintrag/Einträge) — "
                            "Kommandoausführung möglich."
                        ),
                        context={"samples": lines[:3]},
                    )
                )
        except Exception as exc:  # noqa: BLE001 -- olevba/msodde/openpyxl koennen bei defekten Files unspezifizierte Errors werfen, Scanner darf nicht crashen
            _log.debug("msodde scan fehlgeschlagen für %s: %s", path.name, exc)
            report.add(
                Threat(
                    code="XLSX_DDE_SCAN_ERROR",
                    severity=Severity.LOW,
                    message="DDE-Scan konnte nicht abgeschlossen werden.",
                    context={"error": str(exc)[:120]},
                )
            )

    @staticmethod
    def _check_formulas(path: Path, report: ValidationReport) -> None:
        """Durchläuft alle Zellen und meldet Formula-Injection-Muster."""
        try:
            from openpyxl import load_workbook  # noqa: PLC0415
        except ImportError:  # Fallback falls openpyxl nicht verfügbar
            report.add(
                Threat(
                    code="XLSX_FORMULA_SCAN_SKIPPED",
                    severity=Severity.LOW,
                    message="Formula-Scan übersprungen (openpyxl fehlt).",
                )
            )
            return

        try:
            wb = load_workbook(
                filename=str(path),
                read_only=True,
                data_only=False,
                keep_vba=False,
            )
        except Exception as exc:  # noqa: BLE001 -- olevba/msodde/openpyxl koennen bei defekten Files unspezifizierte Errors werfen, Scanner darf nicht crashen
            report.add(
                Threat(
                    code="XLSX_FORMULA_SCAN_ERROR",
                    severity=Severity.LOW,
                    message=f"Formula-Scan konnte nicht starten: {exc}",
                    context={"error": str(exc)[:120]},
                )
            )
            return

        suspicious: list[dict[str, str]] = []
        dangerous_hits: list[dict[str, str]] = []

        try:
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=False):
                    for cell in row:
                        val = cell.value
                        if not isinstance(val, str) or not val:
                            continue
                        first = val[0]
                        if first not in FORMULA_PREFIXES:
                            continue
                        upper = val.upper()
                        is_dangerous = any(fn in upper for fn in DANGEROUS_FORMULAS)
                        sample = {
                            "sheet": sheet.title,
                            "cell": cell.coordinate,
                            # Erste 60 Zeichen — verhindert PII-Logging
                            "formula": val[:60],
                        }
                        if is_dangerous:
                            if len(dangerous_hits) < MAX_FORMULA_SAMPLES:
                                dangerous_hits.append(sample)
                        elif len(suspicious) < MAX_FORMULA_SAMPLES:
                            suspicious.append(sample)
                        # Weiterlaufen, damit Statistik stimmt
        finally:
            wb.close()

        if dangerous_hits:
            report.add(
                Threat(
                    code="XLSX_FORMULA_INJECTION",
                    severity=Severity.HIGH,
                    message=(
                        f"{len(dangerous_hits)} potenziell gefährliche Formel(n) "
                        "erkannt (CMD/DDE/WEBSERVICE/HYPERLINK/IMPORT*)."
                    ),
                    context={"samples": dangerous_hits},
                )
            )
        if suspicious:
            report.add(
                Threat(
                    code="XLSX_FORMULA_SUSPICIOUS",
                    severity=Severity.MEDIUM,
                    message=(
                        f"{len(suspicious)} Zelle(n) mit Formel-Präfix — "
                        "bei CSV-Export als Formula-Injection ausnutzbar."
                    ),
                    context={"samples": suspicious},
                )
            )
