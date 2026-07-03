"""
test_export_layer — Tests fuer den gemeinsamen Export-Layer.

Prueft:
1. BaseExporter: ABC — direkte Instanziierung schlaegt fehl
2. Jeder Tool-Exporter: implementiert das Interface korrekt
3. JSON-Export: erzeugt valide JSON-Datei mit Metadaten
4. XLSX-Export: erzeugt valide.xlsx-Datei mit korrektem Sheet
5. DepExporter / CertExporter: edge cases (leere Listen)

Alle Exports laufen gegen tmp_path (kein echtes Dateisystem).
Keine externen Services — nur Domain-Objekte.

Author: Patrick Riederich
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path

import pytest

from core.export.base_exporter import BaseExporter
from tools.cert_monitor.application.cert_exporter import CertExporter
from tools.cert_monitor.domain.models import CertInfo, CertStatus
from tools.dependency_auditor.application.dep_exporter import DepExporter
from tools.dependency_auditor.domain.models import (
    DependencyAuditResult,
    DependencyInfo,
    VulnerabilityInfo,
    VulnSeverity,
)
from tools.network_scanner.application.network_exporter import NetworkExporter
from tools.network_scanner.domain.models import (
    HostInfo,
    NetworkScanResult,
    PortInfo,
    PortRisk,
    PortState,
)
from tools.system_scanner.application.system_exporter import SystemExporter
from tools.system_scanner.domain.entities import OSInfo, ScanResult, SecurityComponent
from tools.system_scanner.domain.enums import ComponentStatus, ComponentType, OSPlatform

# ---------------------------------------------------------------------------
# Hilfsfunktionen / Fixtures
# ---------------------------------------------------------------------------


def _make_scan_result() -> ScanResult:
    """Erstellt ein minimales ScanResult fuer Tests."""
    return ScanResult(
        scan_id="test-scan-001",
        timestamp=datetime(2026, 4, 8, 12, 0, tzinfo=UTC),
        os_info=OSInfo(
            platform=OSPlatform.WINDOWS,
            name="Windows 11 Home",
            version="10.0.26200",
            build="26200",
            architecture="AMD64",
        ),
        security_components=[
            SecurityComponent(
                name="Windows Defender",
                type=ComponentType.ANTIVIRUS,
                status=ComponentStatus.ACTIVE,
                version="4.18.2403.4",
            )
        ],
        scan_duration_s=1.5,
    )


def _make_network_result() -> NetworkScanResult:
    """Erstellt ein minimales NetworkScanResult fuer Tests."""
    host = HostInfo(
        host="192.168.1.1",
        erreichbar=True,
        offene_ports=[
            PortInfo(
                port=80, state=PortState.OPEN, service="http", risk=PortRisk.MITTEL
            ),
            PortInfo(
                port=443, state=PortState.OPEN, service="https", risk=PortRisk.NIEDRIG
            ),
        ],
    )
    return NetworkScanResult(
        ziel="192.168.1.0/24",
        hosts=[host],
        gestartet_am=datetime(2026, 4, 8, 12, 0, tzinfo=UTC),
        beendet_am=datetime(2026, 4, 8, 12, 1, tzinfo=UTC),
        scanner_typ="socket",
        scan_id="net-test-001",
    )


def _make_certs() -> list[CertInfo]:
    """Erstellt eine minimale CertInfo-Liste fuer Tests."""
    return [
        CertInfo(
            domain="example.com",
            port=443,
            status=CertStatus.OK,
            aussteller="Let's Encrypt",
            gueltig_bis="2026-07-01 00:00:00",
            tage_verbleibend=84,
            tls_version="TLSv1.3",
        ),
        CertInfo(
            domain="old.example.com",
            port=443,
            status=CertStatus.KRITISCH,
            aussteller="Self-Signed",
            gueltig_bis="2026-04-10 00:00:00",
            tage_verbleibend=2,
            tls_version="TLSv1.2",
            ist_self_signed=True,
        ),
    ]


def _make_audit_result() -> DependencyAuditResult:
    """Erstellt ein minimales DependencyAuditResult fuer Tests."""
    vuln = VulnerabilityInfo(
        vuln_id="GHSA-test-0001",
        package_name="requests",
        affected_versions=">=2.0,<2.32",
        fixed_version="2.32.0",
        severity=VulnSeverity.HIGH,
        summary="Test vulnerability in requests",
        url="https://example.com/vuln",
    )
    dep_pinned = DependencyInfo(
        name="requests", version_pinned="2.31.0", version_spec="==2.31.0", line_number=1
    )
    dep_unpinned = DependencyInfo(
        name="flask", version_pinned=None, version_spec=">=2.0", line_number=2
    )
    return DependencyAuditResult(
        source_file="requirements.txt",
        scan_timestamp=datetime(2026, 4, 8, 12, 0, tzinfo=UTC).isoformat(),
        total_dependencies=2,
        total_vulnerabilities=1,
        dependencies=[dep_pinned, dep_unpinned],
        vulnerabilities=[vuln],
        unpinned_dependencies=[dep_unpinned],
        severity_summary={VulnSeverity.HIGH.value: 1},
    )


# ---------------------------------------------------------------------------
# 1. BaseExporter: ABC verhindert direkte Instanziierung
# ---------------------------------------------------------------------------


def test_base_exporter_is_abstract() -> None:
    """BaseExporter kann nicht direkt instanziiert werden."""
    with pytest.raises(TypeError):
        BaseExporter()  # type: ignore[abstract]


# ---------------------------------------------------------------------------
# 2. Alle Exporter: Interface korrekt implementiert
# ---------------------------------------------------------------------------


def test_system_exporter_interface() -> None:
    """SystemExporter implementiert BaseExporter vollstaendig."""
    e = SystemExporter()
    assert isinstance(e, BaseExporter)
    assert e.default_filename_stem == "system_scan_export"


def test_network_exporter_interface() -> None:
    """NetworkExporter implementiert BaseExporter vollstaendig."""
    e = NetworkExporter()
    assert isinstance(e, BaseExporter)
    assert e.default_filename_stem == "netzwerk_scan_export"


def test_cert_exporter_interface() -> None:
    """CertExporter implementiert BaseExporter vollstaendig."""
    e = CertExporter()
    assert isinstance(e, BaseExporter)
    assert e.default_filename_stem == "zertifikate_export"


def test_dep_exporter_interface() -> None:
    """DepExporter implementiert BaseExporter vollstaendig."""
    e = DepExporter()
    assert isinstance(e, BaseExporter)
    assert e.default_filename_stem == "dependency_audit_export"


# ---------------------------------------------------------------------------
# 3. JSON-Export: Validierung der erzeugten Dateien
# ---------------------------------------------------------------------------


def test_system_exporter_json(tmp_path: Path) -> None:
    """SystemExporter.export_json erzeugt valide JSON-Datei."""
    exporter = SystemExporter()
    result = _make_scan_result()
    out = tmp_path / "system_scan.json"

    ok = exporter.export_json(result, str(out))

    assert ok is True
    assert out.exists()
    data = json.loads(out.read_text(encoding="utf-8"))
    assert "meta" in data
    assert data["meta"]["scan_id"] == "test-scan-001"
    assert "os_info" in data
    assert len(data["security_components"]) == 1


def test_network_exporter_json(tmp_path: Path) -> None:
    """NetworkExporter.export_json erzeugt valide JSON-Datei."""
    exporter = NetworkExporter()
    result = _make_network_result()
    out = tmp_path / "network_scan.json"

    ok = exporter.export_json(result, str(out))

    assert ok is True
    assert out.exists()
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["meta"]["ziel"] == "192.168.1.0/24"
    assert data["zusammenfassung"]["erreichbare_hosts"] == 1
    assert len(data["hosts"]) == 1


def test_cert_exporter_json(tmp_path: Path) -> None:
    """CertExporter.export_json erzeugt valide JSON-Datei."""
    exporter = CertExporter()
    certs = _make_certs()
    out = tmp_path / "zertifikate.json"

    ok = exporter.export_json(certs, str(out))

    assert ok is True
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["meta"]["total_count"] == 2
    assert data["zertifikate"][0]["domain"] == "example.com"
    assert data["zertifikate"][1]["ist_self_signed"] is True


def test_dep_exporter_json(tmp_path: Path) -> None:
    """DepExporter.export_json erzeugt valide JSON-Datei."""
    exporter = DepExporter()
    result = _make_audit_result()
    out = tmp_path / "audit.json"

    ok = exporter.export_json(result, str(out))

    assert ok is True
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["meta"]["total_vulnerabilities"] == 1
    assert data["vulnerabilities"][0]["vuln_id"] == "GHSA-test-0001"
    assert len(data["unpinned_dependencies"]) == 1


# ---------------------------------------------------------------------------
# 4. XLSX-Export: openpyxl erzeugt valide Datei
# ---------------------------------------------------------------------------


def test_system_exporter_xlsx(tmp_path: Path) -> None:
    """SystemExporter.export_xlsx erzeugt valide.xlsx-Datei."""
    pytest.importorskip("openpyxl")
    exporter = SystemExporter()
    result = _make_scan_result()
    out = tmp_path / "system_scan.xlsx"

    ok = exporter.export_xlsx(result, str(out))

    assert ok is True
    assert out.exists()
    import openpyxl

    wb = openpyxl.load_workbook(str(out))
    assert "Sicherheitskomponenten" in wb.sheetnames
    assert "OS-Info" in wb.sheetnames
    ws = wb["Sicherheitskomponenten"]
    assert ws.cell(row=1, column=1).value == "Name"


def test_network_exporter_xlsx(tmp_path: Path) -> None:
    """NetworkExporter.export_xlsx erzeugt valide.xlsx-Datei mit Ports."""
    pytest.importorskip("openpyxl")
    exporter = NetworkExporter()
    result = _make_network_result()
    out = tmp_path / "network_scan.xlsx"

    ok = exporter.export_xlsx(result, str(out))

    assert ok is True
    import openpyxl

    wb = openpyxl.load_workbook(str(out))
    assert "Offene Ports" in wb.sheetnames
    assert "Host-Übersicht" in wb.sheetnames


def test_cert_exporter_xlsx(tmp_path: Path) -> None:
    """CertExporter.export_xlsx erzeugt valide.xlsx-Datei."""
    pytest.importorskip("openpyxl")
    exporter = CertExporter()
    certs = _make_certs()
    out = tmp_path / "zertifikate.xlsx"

    ok = exporter.export_xlsx(certs, str(out))

    assert ok is True
    import openpyxl

    wb = openpyxl.load_workbook(str(out))
    assert "Zertifikate" in wb.sheetnames
    ws = wb["Zertifikate"]
    assert ws.max_row == 3  # Header + 2 Zertifikate


def test_dep_exporter_xlsx(tmp_path: Path) -> None:
    """DepExporter.export_xlsx erzeugt valide.xlsx-Datei."""
    pytest.importorskip("openpyxl")
    exporter = DepExporter()
    result = _make_audit_result()
    out = tmp_path / "audit.xlsx"

    ok = exporter.export_xlsx(result, str(out))

    assert ok is True
    import openpyxl

    wb = openpyxl.load_workbook(str(out))
    assert "Vulnerabilities" in wb.sheetnames
    assert "Version unbekannt" in wb.sheetnames
    assert "Unpinned Dependencies" in wb.sheetnames


def test_dep_exporter_xlsx_unverified_sheet(tmp_path: Path) -> None:
    """: 'Version unbekannt'-Sheet enthaelt die unverifizierten Advisories."""
    pytest.importorskip("openpyxl")
    exporter = DepExporter()
    result = _make_audit_result()
    unverified = VulnerabilityInfo(
        vuln_id="GHSA-unverified-01",
        package_name="ghost-package",
        affected_versions="unbekannt",
        fixed_version=None,
        severity=VulnSeverity.MEDIUM,
        summary="Advisory ohne moeglichen Versionsabgleich",
        url="https://example.com/unverified",
    )
    result.unverified_vulnerabilities = [unverified]
    out = tmp_path / "audit_unverified.xlsx"

    ok = exporter.export_xlsx(result, str(out))

    assert ok is True
    import openpyxl

    wb = openpyxl.load_workbook(str(out))
    ws = wb["Version unbekannt"]
    assert ws.cell(row=1, column=1).value == "Advisory-ID"
    assert ws.cell(row=2, column=1).value == "GHSA-unverified-01"
    assert ws.cell(row=2, column=2).value == "ghost-package"
    assert ws.cell(row=2, column=5).value == "—"  # kein Fix bekannt


def test_dep_exporter_pdf_escapet_untrusted_strings(tmp_path: Path) -> None:
    """: OSV-Strings mit Markup crashen den PDF-Export nicht.

    Rohe '<'-Zeichen in Paragraph-Inhalten fuehren bei ReportLab zu
    Parse-Fehlern oder Markup-Injektion — der Export muss sie an der
    Render-Stelle escapen (Choke-Point-Muster wie report_generator).
    """
    pytest.importorskip("reportlab")
    exporter = DepExporter()
    boese_vuln = VulnerabilityInfo(
        vuln_id="GHSA-<b>evil</b>",
        package_name="paket<script>",
        affected_versions="<unbekannt>",
        fixed_version=None,
        severity=VulnSeverity.HIGH,
        summary='<img src="x" onerror="alert(1)"> Boeser Summary',
        url="https://example.com/evil",
    )
    result = _make_audit_result()
    result.vulnerabilities = [boese_vuln]
    result.unverified_vulnerabilities = [boese_vuln]
    result.source_file = "C:/pfad/<kunde>/requirements.txt"
    out = tmp_path / "audit_escaped.pdf"

    ok = exporter.export_pdf(result, str(out))

    assert ok is True
    assert out.exists()
    assert out.stat().st_size > 0


# ---------------------------------------------------------------------------
# 5. Edge cases: leere Daten
# ---------------------------------------------------------------------------


def test_cert_exporter_json_empty(tmp_path: Path) -> None:
    """CertExporter.export_json funktioniert mit leerer Liste."""
    exporter = CertExporter()
    out = tmp_path / "empty.json"

    ok = exporter.export_json([], str(out))

    assert ok is True
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["meta"]["total_count"] == 0
    assert data["zertifikate"] == []


def test_dep_exporter_json_no_vulns(tmp_path: Path) -> None:
    """DepExporter.export_json funktioniert bei Audit ohne Schwachstellen."""
    exporter = DepExporter()
    result = DependencyAuditResult(
        source_file="requirements.txt",
        scan_timestamp=datetime.now(UTC).isoformat(),
        total_dependencies=5,
        total_vulnerabilities=0,
    )
    out = tmp_path / "clean.json"

    ok = exporter.export_json(result, str(out))

    assert ok is True
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["meta"]["total_vulnerabilities"] == 0
    assert data["vulnerabilities"] == []
