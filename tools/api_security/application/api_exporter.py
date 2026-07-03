"""
api_exporter — Export-Service für API-Security-Scanner-Ergebnisse.

Unterstützt JSON, Excel (.xlsx) und PDF (FINLAI Dark Theme).
JSON- und PDF-Export delegiert an ScannerService (bestehende Implementierung).
XLSX-Export ist neu implementiert.

Schichtzugehörigkeit: application/ — kein GUI-Import.

Author: Patrick Riederich
Version: 1.0
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from core.exceptions import ConfigurationError
from core.export.base_exporter import BaseExporter
from core.logger import get_logger

if TYPE_CHECKING:
    from tools.api_security.application.scanner_service import ScannerService
    from tools.api_security.domain.models import ScanResult

log = get_logger(__name__)

# Excel-Farben (ARGB)
_XL_TEAL = "FF26A69A"
_XL_WHITE = "FFFFFFFF"
_XL_ROW_ODD = "FF252525"
_XL_ROW_EVEN = "FF1E1E1E"
_XL_TEXT = "FFC8CCD0"

_SEV_COLORS_ARGB = {
    "critical": "FFD32F2F",
    "high": "FFF57C00",
    "medium": "FFFBC02D",
    "low": "FF388E3C",
    "info": "FF616161",
}


class ApiExporter(BaseExporter):
    """Exportiert ScanResult-Objekte des API-Security-Scanners.

    Attributes:
        _service: ScannerService-Instanz für JSON/PDF-Delegation.
        _report_title: Titel für PDF-Reports.
    """

    _report_title = "API-Security Report"

    def __init__(self, service: ScannerService) -> None:
        """Initialisiert den Exporter.

        Args:
            service: ScannerService mit IReportPort-Abhängigkeit.
        """
        self._service = service

    @property
    def default_filename_stem(self) -> str:
        """Basis-Dateiname.

        Returns:
            String "api_security_export".
        """
        return "api_security_export"

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def export_json(self, data: object, path: str) -> bool:
        """Exportiert ScanResult als JSON via ScannerService.

        Args:
            data: ScanResult-Instanz.
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.
        """
        result: ScanResult = data  # type: ignore[assignment]
        self._service.export_json(result, Path(path))
        return True

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    def export_xlsx(self, data: object, path: str) -> bool:
        """Exportiert ScanResult als Excel-Datei.

        Zwei Sheets: "Findings" + "Zusammenfassung".

        Args:
            data: ScanResult-Instanz.
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.

        Raises:
            RuntimeError: Bei fehlendem openpyxl.
        """
        try:
            import openpyxl  # noqa: PLC0415
            from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415
            from openpyxl.utils import get_column_letter  # noqa: PLC0415
        except ImportError as exc:
            raise ConfigurationError("openpyxl nicht installiert") from exc

        result: ScanResult = data  # type: ignore[assignment]
        wb = openpyxl.Workbook()

        header_fill = PatternFill("solid", fgColor=_XL_TEAL)
        header_font = Font(name="Calibri", bold=True, color=_XL_WHITE)
        odd_fill = PatternFill("solid", fgColor=_XL_ROW_ODD)
        even_fill = PatternFill("solid", fgColor=_XL_ROW_EVEN)
        text_font = Font(name="Calibri", color=_XL_TEXT)

        # Sheet 1: Findings
        ws1 = wb.active
        ws1.title = "Findings"  # type: ignore[union-attr]
        headers1 = ["Code", "Titel", "Schweregrad", "OWASP", "Beschreibung", "Maßnahme"]
        for col_idx, h in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)  # type: ignore[union-attr]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        findings = result.findings_by_severity()
        for row_idx, finding in enumerate(findings, 2):
            fill = odd_fill if row_idx % 2 == 1 else even_fill
            sev_argb = _SEV_COLORS_ARGB.get(finding.severity.value, _XL_TEXT)
            sev_font = Font(name="Calibri", color=sev_argb)
            values = [
                finding.code,
                finding.title,
                finding.severity.label(),
                f"{finding.owasp.value}: {finding.owasp.description()}",
                finding.description[:200],
                finding.remediation[:150] if finding.remediation else "—",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)  # type: ignore[union-attr]
                cell.fill = fill
                cell.font = sev_font if col_idx == 3 else text_font

        for col_idx, width in enumerate([20, 35, 12, 40, 60, 50], 1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]
        ws1.freeze_panes = "A2"  # type: ignore[union-attr]

        # Sheet 2: Zusammenfassung
        ws2 = wb.create_sheet("Zusammenfassung")
        summary_rows = [
            ("Ziel-URL", result.target.url),
            ("API-Typ", result.target.api_type.value),
            ("Scan-Zeitstempel", result.scan_time),
            ("Dauer (ms)", str(result.duration_ms)),
            ("Gesamt Findings", str(len(result.findings))),
            ("Kritisch", str(result.critical_count())),
            ("Hoch", str(result.high_count())),
            ("Risiko-Score", str(result.risk_score())),
        ]
        for i, (key, val) in enumerate(summary_rows, 1):
            k_cell = ws2.cell(row=i, column=1, value=key)
            k_cell.font = Font(name="Calibri", bold=True, color=_XL_TEAL)
            v_cell = ws2.cell(row=i, column=2, value=val)
            v_cell.font = text_font
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 50

        wb.save(path)
        log.info(
            "API-Security XLSX-Export: %d Findings → %s", len(result.findings), path
        )
        return True

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    def export_pdf(
        self,
        data: object,
        path: str,
        title: str = "",
        subtitle: str = "",
    ) -> bool:
        """Exportiert ScanResult als PDF via ScannerService.

        Args:
            data: ScanResult-Instanz.
            path: Zieldateipfad.
            title: Nicht verwendet (Titel stammt aus ScannerService-Reporter).
            subtitle: Nicht verwendet.

        Returns:
            True bei Erfolg.
        """
        result: ScanResult = data  # type: ignore[assignment]
        self._service.export_pdf(result, Path(path))
        return True
