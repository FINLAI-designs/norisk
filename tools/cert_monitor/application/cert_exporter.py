"""
cert_exporter — Export-Service für Zertifikats-Monitor-Ergebnisse.

Unterstützt JSON, Excel (.xlsx) und PDF (FINLAI Dark Theme).

Schichtzugehörigkeit: application/ — kein GUI-Import.

Author: Patrick Riederich
Version: 1.0
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from core import theme
from core.exceptions import ConfigurationError
from core.export.base_exporter import BaseExporter
from core.logger import get_logger

if TYPE_CHECKING:
    from tools.cert_monitor.domain.models import CertInfo

log = get_logger(__name__)

# Excel-Farben (ARGB)
_XL_TEAL = "FF26A69A"
_XL_WHITE = "FFFFFFFF"
_XL_ROW_ODD = "FF252525"
_XL_ROW_EVEN = "FF1E1E1E"
_XL_TEXT = "FFC8CCD0"

_STATUS_DE = {
    "ok": "OK",
    "warnung": "Warnung",
    "kritisch": "Kritisch",
    "fehler": "Fehler",
    "unbekannt": "Unbekannt",
}

# Hinweis: "unbekannt" verwendet literal "#606070" (sichtbar dunkler als
# theme.SEVERITY_SIGNAL_INFO="#888888"); bewusste Domäne-Wahl für CertStatus.UNBEKANNT.
_STATUS_COLORS_PDF = {
    "ok": theme.SEVERITY_SIGNAL_OK,
    "warnung": theme.SEVERITY_SIGNAL_MEDIUM,
    "kritisch": theme.SEVERITY_SIGNAL_CRITICAL,
    "fehler": theme.SEVERITY_SIGNAL_INFO,
    "unbekannt": "#606070",  # noqa: domain-cert-status-unknown
}


class CertExporter(BaseExporter):
    """Exportiert CertInfo-Listen in JSON, XLSX und PDF.

    Attributes:
        _report_title: Titel für PDF-Reports.
    """

    _report_title = "Zertifikats-Monitor Report"

    @property
    def default_filename_stem(self) -> str:
        """Basis-Dateiname.

        Returns:
            String "zertifikate_export".
        """
        return "zertifikate_export"

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def export_json(self, data: object, path: str) -> bool:
        """Exportiert CertInfo-Liste als JSON.

        Args:
            data: list[CertInfo].
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.
        """
        certs: list[CertInfo] = data  # type: ignore[assignment]
        export_dt = datetime.now(UTC).isoformat()

        status_counts: dict[str, int] = {}
        for cert in certs:
            key = cert.status.value
            status_counts[key] = status_counts.get(key, 0) + 1

        payload = {
            "meta": {
                "exported_at": export_dt,
                "generator": "NoRisk by FINLAI — Zertifikats-Monitor Export",
                "total_count": len(certs),
                "status_summary": status_counts,
            },
            "zertifikate": [
                {
                    "domain": cert.domain,
                    "port": cert.port,
                    "status": cert.status.value,
                    "aussteller": cert.aussteller,
                    "gueltig_von": cert.gueltig_von,
                    "gueltig_bis": cert.gueltig_bis,
                    "tage_verbleibend": cert.tage_verbleibend,
                    "tls_version": cert.tls_version,
                    "cipher_name": cert.cipher_name,
                    "cipher_bits": cert.cipher_bits,
                    "ist_self_signed": cert.ist_self_signed,
                    "san_domains": cert.san_domains,
                    "serial_number": cert.serial_number,
                    "findings": cert.findings,
                    "letzte_pruefung": cert.letzte_pruefung,
                    "fehler_meldung": cert.fehler_meldung,
                }
                for cert in certs
            ],
        }
        Path(path).write_text(
            json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        log.info(
            "Zertifikats-Monitor JSON-Export: %d Zertifikate → %s", len(certs), path
        )
        return True

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    def export_xlsx(self, data: object, path: str) -> bool:
        """Exportiert CertInfo-Liste als Excel-Datei.

        Args:
            data: list[CertInfo].
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.

        Raises:
            RuntimeError: Bei fehlendem openpyxl.
        """
        try:
            import openpyxl  # noqa: PLC0415
            from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415
            from openpyxl.utils import get_column_letter  # noqa: PLC0415
        except ImportError as exc:
            raise ConfigurationError("openpyxl nicht installiert") from exc

        certs: list[CertInfo] = data  # type: ignore[assignment]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Zertifikate"  # type: ignore[union-attr]

        header_fill = PatternFill("solid", fgColor=_XL_TEAL)
        header_font = Font(name="Calibri", bold=True, color=_XL_WHITE)
        odd_fill = PatternFill("solid", fgColor=_XL_ROW_ODD)
        even_fill = PatternFill("solid", fgColor=_XL_ROW_EVEN)
        text_font = Font(name="Calibri", color=_XL_TEXT)

        headers = [
            "Domain",
            "Status",
            "Aussteller",
            "Gültig von",
            "Gültig bis",
            "Tage verbleibend",
            "TLS-Version",
            "Cipher",
            "Bits",
            "Self-Signed",
            "Findings",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)  # type: ignore[union-attr]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, cert in enumerate(certs, 2):
            fill = odd_fill if row_idx % 2 == 1 else even_fill
            values = [
                cert.anzeige_domain,
                _STATUS_DE.get(cert.status.value, cert.status.value),
                cert.aussteller or "—",
                cert.gueltig_von or "—",
                cert.gueltig_bis or "—",
                cert.tage_verbleibend,
                cert.tls_version or "—",
                cert.cipher_name or "—",
                cert.cipher_bits or "—",
                "Ja" if cert.ist_self_signed else "Nein",
                "; ".join(cert.findings) if cert.findings else "—",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)  # type: ignore[union-attr]
                cell.fill = fill
                cell.font = text_font

        col_widths = [30, 12, 30, 18, 18, 18, 12, 30, 8, 12, 50]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]
        ws.freeze_panes = "A2"  # type: ignore[union-attr]

        wb.save(path)
        log.info(
            "Zertifikats-Monitor XLSX-Export: %d Zertifikate → %s", len(certs), path
        )
        return True

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    def export_pdf(
        self,
        data: object,
        path: str,
        title: str = "",
        subtitle: str = "",
    ) -> bool:
        """Exportiert CertInfo-Liste als Dark-Theme PDF.

        Args:
            data: list[CertInfo].
            path: Zieldateipfad.
            title: Optionaler Titel.
            subtitle: Optionaler Untertitel.

        Returns:
            True bei Erfolg.
        """
        from reportlab.lib.units import cm  # noqa: PLC0415
        from reportlab.platypus import (  # noqa: PLC0415
            Paragraph,
            Spacer,
            Table,
            TableStyle,
        )

        from core.pdf.pdf_colors import (  # noqa: PLC0415
            PDF_BG_PAGE,
            PDF_TABLE_HEADER_BG,
            PDF_TABLE_HEADER_TEXT,
            PDF_TABLE_ROW_EVEN,
            PDF_TABLE_ROW_ODD,
            PDF_TEXT_PRIMARY,
        )
        from core.pdf.pdf_fonts import FONT_RALEWAY, FONT_RALEWAY_BOLD  # noqa: PLC0415
        from core.pdf.pdf_report_builder import DarkReportBuilder  # noqa: PLC0415

        certs: list[CertInfo] = data  # type: ignore[assignment]
        export_dt = datetime.now(UTC).strftime("%d.%m.%Y %H:%M UTC")

        _title = title or self._report_title
        _subtitle = subtitle or f"NoRisk by FINLAI  ·  {export_dt}"

        status_counts: dict[str, int] = {}
        for cert in certs:
            key = cert.status.value
            status_counts[key] = status_counts.get(key, 0) + 1

        builder = DarkReportBuilder(output_path=path, title=_title, subtitle=_subtitle)
        builder.add_cover(date_str=export_dt)

        st = builder._styles  # noqa: SLF001
        story = builder._story  # noqa: SLF001

        # Zusammenfassung
        story.append(Paragraph("Zertifikats-Übersicht", st["h2"]))
        story.append(
            Paragraph(
                f"Gesamt: {len(certs)}  |  "
                f"OK: {status_counts.get('ok', 0)}  |  "
                f"Warnung: {status_counts.get('warnung', 0)}  |  "
                f"Kritisch: {status_counts.get('kritisch', 0)}  |  "
                f"Fehler: {status_counts.get('fehler', 0)}",
                st["body_dim"],
            )
        )
        story.append(Spacer(1, 0.3 * cm))

        # Zertifikats-Tabelle
        col_widths = [5 * cm, 2.5 * cm, 4.5 * cm, 2.5 * cm, 3 * cm]
        header_row = ["Domain", "Status", "Aussteller", "Gültig bis", "TLS"]
        table_data: list = [[Paragraph(h, st["table_header"]) for h in header_row]]

        for cert in certs:
            scolor = _STATUS_COLORS_PDF.get(
                cert.status.value, theme.SEVERITY_SIGNAL_INFO
            )
            status_label = _STATUS_DE.get(cert.status.value, cert.status.value)

            # Datum formatieren
            gueltig_bis_anzeige = cert.gueltig_bis
            if cert.gueltig_bis:
                for _fmt in (
                    "%b %d %H:%M:%S %Y %Z",
                    "%b  %d %H:%M:%S %Y %Z",
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d",
                ):
                    try:
                        from datetime import datetime as _dt  # noqa: PLC0415

                        _d = _dt.strptime(cert.gueltig_bis, _fmt)
                        gueltig_bis_anzeige = _d.strftime("%d.%m.%Y")
                        break
                    except ValueError:
                        continue

            table_data.append(
                [
                    Paragraph(cert.anzeige_domain, st["table_cell"]),
                    Paragraph(
                        f'<font color="{scolor}">{status_label}</font>',
                        st["table_cell"],
                    ),
                    Paragraph(
                        cert.aussteller[:35] if cert.aussteller else "—",
                        st["table_cell"],
                    ),
                    Paragraph(gueltig_bis_anzeige, st["table_cell"]),
                    Paragraph(cert.tls_version or "—", st["table_cell"]),
                ]
            )

        if len(table_data) > 1:
            row_count = len(table_data)
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), PDF_TABLE_HEADER_BG),
                        ("TEXTCOLOR", (0, 0), (-1, 0), PDF_TABLE_HEADER_TEXT),
                        ("FONTNAME", (0, 0), (-1, 0), FONT_RALEWAY_BOLD),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 0), (-1, 0), 8),
                        ("FONTNAME", (0, 1), (-1, -1), FONT_RALEWAY),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, row_count - 1),
                            [PDF_TABLE_ROW_ODD, PDF_TABLE_ROW_EVEN],
                        ),
                        ("GRID", (0, 0), (-1, -1), 0.3, PDF_BG_PAGE),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 1), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                        ("TEXTCOLOR", (0, 1), (-1, -1), PDF_TEXT_PRIMARY),
                    ]
                )
            )
            story.append(tbl)

        builder.add_footer_page()
        builder.build()
        log.info(
            "Zertifikats-Monitor PDF-Export: %d Zertifikate → %s", len(certs), path
        )
        return True
