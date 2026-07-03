"""
csaf_exporter — Export-Service für CSAF Advisory-Listen.

Unterstützt drei Formate:
  - Excel (.xlsx) via openpyxl — Teal-Header, abwechselnde Zeilenfarben
  - JSON (.json) — pretty-printed mit Metadaten
  - PDF (.pdf) — FINLAI Dark Theme via core/pdf/DarkReportBuilder

Schichtzugehörigkeit: application/ — kein GUI-Import.

Author: Patrick Riederich
Version: 1.0
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from core.exceptions import ConfigurationError
from core.logger import get_logger

if TYPE_CHECKING:
    from tools.csaf_advisor.domain.advisory import CsafAdvisory

log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Schweregrad-Übersetzung
# ---------------------------------------------------------------------------
_SEV_DE: dict[str, str] = {
    "critical": "KRITISCH",
    "high": "HOCH",
    "medium": "MITTEL",
    "low": "NIEDRIG",
}

# Excel-Farben (ARGB)
_XL_TEAL = "FF26A69A"
_XL_WHITE = "FFFFFFFF"
_XL_ROW_ODD = "FF1E2A35"
_XL_ROW_EVEN = "FF141414"
_XL_TEXT = "FFC8CCD0"


def _sev_de(severity: str) -> str:
    return _SEV_DE.get(severity.lower(), severity.upper())


# ---------------------------------------------------------------------------
# Excel-Export
# ---------------------------------------------------------------------------


def export_excel(advisories: list[CsafAdvisory], path: str) -> None:
    """Exportiert Advisories als Excel-Datei (.xlsx).

    Teal-Header, abwechselnde Zeilen, auto-Spaltenbreite.

    Args:
        advisories: Liste der zu exportierenden Advisories.
        path: Zieldateipfad.

    Raises:
        RuntimeError: Bei Schreibfehler oder fehlendem openpyxl.
    """
    try:
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415
        from openpyxl.utils import get_column_letter  # noqa: PLC0415
    except ImportError as exc:
        raise ConfigurationError("openpyxl nicht installiert") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CSAF Advisories"  # type: ignore[union-attr]

    headers = [
        "Advisory-ID",
        "Titel",
        "Herausgeber",
        "Schweregrad",
        "CVSS",
        "CVE-IDs",
        "Veröffentlicht",
        "Aktualisiert",
    ]

    header_fill = PatternFill("solid", fgColor=_XL_TEAL)
    header_font = Font(name="Calibri", bold=True, color=_XL_WHITE)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)  # type: ignore[union-attr]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    odd_fill = PatternFill("solid", fgColor=_XL_ROW_ODD)
    even_fill = PatternFill("solid", fgColor=_XL_ROW_EVEN)
    text_font = Font(name="Calibri", color=_XL_TEXT)

    for row_idx, adv in enumerate(advisories, start=2):
        fill = odd_fill if row_idx % 2 == 1 else even_fill
        values = [
            adv.tracking_id,
            adv.title,
            adv.publisher,
            _sev_de(adv.severity),
            f"{adv.cvss_score:.1f}" if adv.cvss_score is not None else "—",
            ", ".join(adv.cve_ids) if adv.cve_ids else "—",
            adv.initial_release or "—",
            adv.current_release or "—",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)  # type: ignore[union-attr]
            cell.fill = fill
            cell.font = text_font

    col_widths = [18, 60, 20, 12, 8, 30, 14, 14]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]

    ws.freeze_panes = "A2"  # type: ignore[union-attr]
    wb.save(path)
    log.info("Excel-Export: %d Advisories → %s", len(advisories), path)


# ---------------------------------------------------------------------------
# JSON-Export
# ---------------------------------------------------------------------------


def export_json(
    advisories: list[CsafAdvisory],
    path: str,
    filter_info: dict | None = None,
) -> None:
    """Exportiert Advisories als JSON-Datei mit Metadaten.

    Args:
        advisories: Liste der zu exportierenden Advisories.
        path: Zieldateipfad.
        filter_info: Optionale Metadaten zu den angewandten Filtern.

    Raises:
        RuntimeError: Bei Schreibfehler.
    """
    export_dt = datetime.now(UTC).isoformat()

    severity_counts: dict[str, int] = {}
    for adv in advisories:
        sev = adv.severity.lower()
        severity_counts[sev] = severity_counts.get(sev, 0) + 1

    payload = {
        "meta": {
            "exported_at": export_dt,
            "total_count": len(advisories),
            "severity_summary": severity_counts,
            "filters": filter_info or {},
            "generator": "NoRisk by FINLAI — CSAF Advisory Export",
        },
        "advisories": [
            {
                "id": adv.id,
                "tracking_id": adv.tracking_id,
                "title": adv.title,
                "publisher": adv.publisher,
                "severity": adv.severity,
                "cvss_score": adv.cvss_score,
                "cve_ids": adv.cve_ids,
                "affected_products": adv.affected_products,
                "initial_release": adv.initial_release,
                "current_release": adv.current_release,
                "summary": adv.summary,
                "source_url": adv.source_url,
                "fetched_at": adv.fetched_at,
            }
            for adv in advisories
        ],
    }

    Path(path).write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    log.info("JSON-Export: %d Advisories → %s", len(advisories), path)


# ---------------------------------------------------------------------------
# PDF-Export
# ---------------------------------------------------------------------------

_SEV_RISK_MAP: dict[str, str] = {
    "critical": "Kritisch",
    "high": "Hoch",
    "medium": "Mittel",
    "low": "Niedrig",
}


def export_pdf(advisories: list[CsafAdvisory], path: str) -> None:
    """Exportiert Advisories als Dark-Theme PDF-Report.

    Nutzt DarkReportBuilder für konsistentes FINLAI-Design.

    Args:
        advisories: Liste der zu exportierenden Advisories.
        path: Zieldateipfad.

    Raises:
        RuntimeError: Bei Fehler während der PDF-Erstellung.
    """
    from reportlab.lib.units import cm  # noqa: PLC0415
    from reportlab.platypus import (  # noqa: PLC0415
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    from core.pdf.pdf_colors import (  # noqa: PLC0415
        PDF_BG_PAGE,
        PDF_TABLE_HEADER_BG,
        PDF_TABLE_HEADER_TEXT,
        PDF_TABLE_ROW_EVEN,
        PDF_TABLE_ROW_ODD,
        PDF_TEXT_PRIMARY,
        risk_color,
    )
    from core.pdf.pdf_fonts import FONT_RALEWAY, FONT_RALEWAY_BOLD  # noqa: PLC0415
    from core.pdf.pdf_report_builder import DarkReportBuilder  # noqa: PLC0415

    export_dt = datetime.now(UTC).strftime("%d.%m.%Y %H:%M UTC")

    severity_counts: dict[str, int] = {}
    for adv in advisories:
        sev = adv.severity.lower()
        severity_counts[sev] = severity_counts.get(sev, 0) + 1

    builder = DarkReportBuilder(
        output_path=path,
        title="CSAF Advisory Report",
        subtitle=f"NoRisk by FINLAI  ·  {export_dt}",
    )
    builder.add_cover(date_str=export_dt)

    st = builder._styles  # noqa: SLF001
    story = builder._story  # noqa: SLF001

    # Kein zusätzlicher PageBreak — add_cover fügt bereits einen ein.
    story.append(Paragraph("Advisory-Übersicht", st["h2"]))
    story.append(
        Paragraph(
            f"Gesamt: {len(advisories)} Advisories  |  "
            f"Kritisch: {severity_counts.get('critical', 0)}  "
            f"Hoch: {severity_counts.get('high', 0)}  "
            f"Mittel: {severity_counts.get('medium', 0)}  "
            f"Niedrig: {severity_counts.get('low', 0)}",
            st["body_dim"],
        )
    )
    story.append(Spacer(1, 0.3 * cm))

    # Tabelle aufbauen
    col_widths = [3.5 * cm, 7.5 * cm, 2.5 * cm, 1.5 * cm, 2.5 * cm]
    header_row = ["Advisory-ID", "Titel", "Herausgeber", "CVSS", "Veröffentlicht"]
    table_data: list = [[Paragraph(h, st["table_header"]) for h in header_row]]

    for adv in advisories:
        risk = _SEV_RISK_MAP.get(adv.severity.lower(), "Mittel")
        sev_color = risk_color(risk)
        r = int(sev_color.red * 255)
        g = int(sev_color.green * 255)
        b = int(sev_color.blue * 255)
        hex_color = f"#{r:02x}{g:02x}{b:02x}"

        cvss_text = f"{adv.cvss_score:.1f}" if adv.cvss_score is not None else "—"
        title_short = adv.title[:75] + "…" if len(adv.title) > 75 else adv.title
        rel_date = adv.initial_release[:10] if adv.initial_release else "—"

        table_data.append(
            [
                Paragraph(
                    f'<font color="{hex_color}">{adv.tracking_id}</font>',
                    st["table_cell"],
                ),
                Paragraph(title_short, st["table_cell"]),
                Paragraph(adv.publisher[:20], st["table_cell"]),
                Paragraph(cvss_text, st["table_cell_center"]),
                Paragraph(rel_date, st["table_cell"]),
            ]
        )

    row_count = len(table_data)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), PDF_TABLE_HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), PDF_TABLE_HEADER_TEXT),
        ("FONTNAME", (0, 0), (-1, 0), FONT_RALEWAY_BOLD),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), FONT_RALEWAY),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        (
            "ROWBACKGROUNDS",
            (0, 1),
            (-1, row_count - 1),
            [PDF_TABLE_ROW_ODD, PDF_TABLE_ROW_EVEN],
        ),
        ("GRID", (0, 0), (-1, -1), 0.3, PDF_BG_PAGE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("TEXTCOLOR", (0, 1), (-1, -1), PDF_TEXT_PRIMARY),
    ]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)

    # Disclaimer
    story.append(Spacer(1, 0.6 * cm))
    story.append(
        Paragraph(
            "Dieser Report wurde automatisch erstellt und dient ausschließlich "
            "internen Informationszwecken. CSAF-Daten stammen von den konfigurierten "
            "Trusted Providern. Alle Angaben ohne Gewähr.",
            st["disclaimer"],
        )
    )

    builder.build()
    log.info("PDF-Export: %d Advisories → %s", len(advisories), path)
