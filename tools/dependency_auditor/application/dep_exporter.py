"""
dep_exporter — Export-Service für Dependency-Auditor-Ergebnisse.

Unterstützt JSON, Excel (.xlsx) und PDF (FINLAI Dark Theme).

Schichtzugehörigkeit: application/ — kein GUI-Import.

Author: Patrick Riederich
Version: 1.0
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from core import theme
from core.escape import escape_html
from core.exceptions import ConfigurationError
from core.export.base_exporter import BaseExporter
from core.logger import get_logger

if TYPE_CHECKING:
    from tools.dependency_auditor.domain.models import DependencyAuditResult

log = get_logger(__name__)

# Excel-Farben (ARGB)
_XL_TEAL = "FF26A69A"
_XL_WHITE = "FFFFFFFF"
_XL_ROW_ODD = "FF252525"
_XL_ROW_EVEN = "FF1E1E1E"
_XL_TEXT = "FFC8CCD0"

# Excel-ARGB ist ein 8-stelliges Format ("FF" + 6-stelliges Hex). Daher behalten
# wir die Excel-Konstanten literal, leiten sie aber von den theme-Severity-Deep-
# Farben ab — Single Source of Truth bleibt theme.py.
_SEV_COLORS_ARGB = {
    "KRITISCH": "FF" + theme.SEVERITY_DEEP_CRITICAL.lstrip("#").upper(),
    "HOCH": "FF" + theme.SEVERITY_DEEP_HIGH.lstrip("#").upper(),
    "MITTEL": "FF" + theme.SEVERITY_DEEP_MEDIUM.lstrip("#").upper(),
    "NIEDRIG": "FF" + theme.SEVERITY_DEEP_LOW.lstrip("#").upper(),
}

_SEV_COLORS_PDF = {
    "KRITISCH": theme.SEVERITY_DEEP_CRITICAL,
    "HOCH": theme.SEVERITY_DEEP_HIGH,
    "MITTEL": theme.SEVERITY_DEEP_MEDIUM,
    "NIEDRIG": theme.SEVERITY_DEEP_LOW,
}

# Maximale Listen-Eintraege in PDF-Textabschnitten (Unpinned / Unverified).
_PDF_MAX_LIST_ITEMS = 30


class DepExporter(BaseExporter):
    """Exportiert DependencyAuditResult-Objekte in JSON, XLSX und PDF.

    Attributes:
        _report_title: Titel für PDF-Reports.
    """

    _report_title = "Dependency-Audit Report"

    @property
    def default_filename_stem(self) -> str:
        """Basis-Dateiname.

        Returns:
            String "dependency_audit_export".
        """
        return "dependency_audit_export"

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def export_json(self, data: object, path: str) -> bool:
        """Exportiert DependencyAuditResult als JSON.

        Args:
            data: DependencyAuditResult-Instanz.
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.
        """
        result: DependencyAuditResult = data  # type: ignore[assignment]
        export_dt = datetime.now(UTC).isoformat()

        payload = {
            "meta": {
                "exported_at": export_dt,
                "generator": "NoRisk by FINLAI — Dependency-Audit Export",
                "source_file": result.source_file,
                "scan_timestamp": result.scan_timestamp,
                "total_dependencies": result.total_dependencies,
                "total_vulnerabilities": result.total_vulnerabilities,
                "total_unverified": result.unverified_count(),
                "severity_summary": result.severity_summary,
            },
            "vulnerabilities": [
                {
                    "vuln_id": v.vuln_id,
                    "package_name": v.package_name,
                    "severity": v.severity.value,
                    "affected_versions": v.affected_versions,
                    "fixed_version": v.fixed_version,
                    "summary": v.summary,
                    "url": v.url,
                }
                for v in sorted(
                    result.vulnerabilities, key=lambda x: x.severity.sort_order()
                )
            ],
            # Advisories ohne moeglichen Versionsabgleich — getrennt
            # von den verifizierten Vulnerabilities exportiert.
            "unverified_vulnerabilities": [
                {
                    "vuln_id": v.vuln_id,
                    "package_name": v.package_name,
                    "severity": v.severity.value,
                    "affected_versions": v.affected_versions,
                    "fixed_version": v.fixed_version,
                    "summary": v.summary,
                    "url": v.url,
                }
                for v in result.unverified_vulnerabilities
            ],
            "unpinned_dependencies": [
                {
                    "name": d.name,
                    "version_spec": d.version_spec,
                    "line_number": d.line_number,
                }
                for d in result.unpinned_dependencies
            ],
        }
        Path(path).write_text(
            json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        log.info(
            "Dependency-Audit JSON-Export: %d Vulns → %s",
            len(result.vulnerabilities),
            path,
        )
        return True

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    def export_xlsx(self, data: object, path: str) -> bool:
        """Exportiert DependencyAuditResult als Excel-Datei.

        Drei Sheets: "Vulnerabilities" + "Version unbekannt" +
        "Unpinned Dependencies".

        Args:
            data: DependencyAuditResult-Instanz.
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.

        Raises:
            RuntimeError: Bei fehlendem openpyxl.
        """
        try:
            import openpyxl  # noqa: PLC0415
            from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415
            from openpyxl.utils import get_column_letter  # noqa: PLC0415
        except ImportError as exc:
            raise ConfigurationError("openpyxl nicht installiert") from exc

        result: DependencyAuditResult = data  # type: ignore[assignment]
        wb = openpyxl.Workbook()

        header_fill = PatternFill("solid", fgColor=_XL_TEAL)
        header_font = Font(name="Calibri", bold=True, color=_XL_WHITE)
        odd_fill = PatternFill("solid", fgColor=_XL_ROW_ODD)
        even_fill = PatternFill("solid", fgColor=_XL_ROW_EVEN)
        text_font = Font(name="Calibri", color=_XL_TEXT)

        # Sheet 1: Vulnerabilities
        ws1 = wb.active
        ws1.title = "Vulnerabilities"  # type: ignore[union-attr]
        headers1 = [
            "Advisory-ID",
            "Package",
            "Schweregrad",
            "Betroffene Versionen",
            "Fix-Version",
            "Zusammenfassung",
        ]
        for col_idx, h in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)  # type: ignore[union-attr]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sorted_vulns = sorted(
            result.vulnerabilities, key=lambda x: x.severity.sort_order()
        )
        for row_idx, vuln in enumerate(sorted_vulns, 2):
            fill = odd_fill if row_idx % 2 == 1 else even_fill
            sev_color = _SEV_COLORS_ARGB.get(vuln.severity.value, _XL_TEXT)
            sev_font = Font(name="Calibri", color=sev_color)
            values = [
                vuln.vuln_id,
                vuln.package_name,
                vuln.severity.value,
                vuln.affected_versions,
                vuln.fixed_version or "—",
                vuln.summary[:200],
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)  # type: ignore[union-attr]
                cell.fill = fill
                cell.font = sev_font if col_idx == 3 else text_font

        for col_idx, width in enumerate([30, 25, 12, 25, 15, 80], 1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]
        ws1.freeze_panes = "A2"  # type: ignore[union-attr]

        # Sheet 2: Advisories ohne moeglichen Versionsabgleich —
        # analog zum JSON-Block "unverified_vulnerabilities", sonst ginge
        # diese Kategorie im Excel-Export still verloren.
        ws_unverified = wb.create_sheet("Version unbekannt")
        for col_idx, h in enumerate(headers1, 1):
            cell = ws_unverified.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, vuln in enumerate(result.unverified_vulnerabilities, 2):
            fill = odd_fill if row_idx % 2 == 1 else even_fill
            values = [
                vuln.vuln_id,
                vuln.package_name,
                vuln.severity.value,
                vuln.affected_versions,
                vuln.fixed_version or "—",
                vuln.summary[:200],
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws_unverified.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.font = text_font

        for col_idx, width in enumerate([30, 25, 12, 25, 15, 80], 1):
            ws_unverified.column_dimensions[get_column_letter(col_idx)].width = width
        ws_unverified.freeze_panes = "A2"

        # Sheet 3: Unpinned Dependencies
        ws2 = wb.create_sheet("Unpinned Dependencies")
        headers2 = ["Package", "Versions-Spezifikation", "Zeile"]
        for col_idx, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, dep in enumerate(result.unpinned_dependencies, 2):
            fill = odd_fill if row_idx % 2 == 1 else even_fill
            for col_idx, val in enumerate(
                [dep.name, dep.version_spec or "—", dep.line_number], 1
            ):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.font = text_font

        for col_idx, width in enumerate([30, 30, 8], 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(path)
        log.info(
            "Dependency-Audit XLSX-Export: %d Vulns → %s",
            len(result.vulnerabilities),
            path,
        )
        return True

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    def export_pdf(
        self,
        data: object,
        path: str,
        title: str = "",
        subtitle: str = "",
    ) -> bool:
        """Exportiert DependencyAuditResult als Dark-Theme PDF.

        Args:
            data: DependencyAuditResult-Instanz.
            path: Zieldateipfad.
            title: Optionaler Titel.
            subtitle: Optionaler Untertitel.

        Returns:
            True bei Erfolg.
        """
        from reportlab.lib.units import cm  # noqa: PLC0415
        from reportlab.platypus import (  # noqa: PLC0415
            Paragraph,
            Spacer,
            Table,
            TableStyle,
        )

        from core.pdf.pdf_colors import (  # noqa: PLC0415
            PDF_BG_PAGE,
            PDF_TABLE_HEADER_BG,
            PDF_TABLE_HEADER_TEXT,
            PDF_TABLE_ROW_EVEN,
            PDF_TABLE_ROW_ODD,
            PDF_TEXT_PRIMARY,
        )
        from core.pdf.pdf_fonts import FONT_RALEWAY, FONT_RALEWAY_BOLD  # noqa: PLC0415
        from core.pdf.pdf_report_builder import DarkReportBuilder  # noqa: PLC0415

        result: DependencyAuditResult = data  # type: ignore[assignment]
        export_dt = datetime.now(UTC).strftime("%d.%m.%Y %H:%M UTC")

        _title = title or self._report_title
        _subtitle = subtitle or f"NoRisk by FINLAI  ·  {export_dt}"

        builder = DarkReportBuilder(output_path=path, title=_title, subtitle=_subtitle)
        builder.add_cover(date_str=export_dt)

        st = builder._styles  # noqa: SLF001
        story = builder._story  # noqa: SLF001

        # Zusammenfassung — source_file ist ein User-gewaehlter Pfad und
        # landet in einem Paragraph (XML-Kontext) → escapen.
        story.append(Paragraph("Audit-Zusammenfassung", st["h2"]))
        story.append(
            Paragraph(
                f"Quelle: {escape_html(result.source_file)}  |  "
                f"Dependencies: {result.total_dependencies}  |  "
                f"Schwachstellen: {result.total_vulnerabilities}  |  "
                f"Kritisch: {result.critical_count()}  |  "
                f"Hoch: {result.high_count()}  |  "
                f"Version unbekannt: {result.unverified_count()}",
                st["body_dim"],
            )
        )
        story.append(Spacer(1, 0.3 * cm))

        # Vulnerability-Tabelle
        story.append(Paragraph("Gefundene Schwachstellen", st["h2"]))

        col_widths = [3.5 * cm, 3 * cm, 2.5 * cm, 3.5 * cm, 5 * cm]
        header_row = [
            "Advisory-ID",
            "Package",
            "Schweregrad",
            "Betroffene Vers.",
            "Zusammenfassung",
        ]
        table_data: list = [[Paragraph(h, st["table_header"]) for h in header_row]]

        sorted_vulns = sorted(
            result.vulnerabilities, key=lambda x: x.severity.sort_order()
        )
        # (escape-at-render): vuln_id/package/affected/summary
        # stammen aus der OSV-Antwort (untrusted) und werden als Paragraph
        # (XML-Kontext) gerendert → durch den Escape-Choke-Point schicken.
        # severity.value ist ein interner Enum-Wert, sev_color Theme —
        # das <font>-Markup bleibt bewusst roh.
        for vuln in sorted_vulns:
            sev_color = _SEV_COLORS_PDF.get(
                vuln.severity.value, theme.SEVERITY_SIGNAL_INFO
            )
            summary_short = vuln.summary[:100] + (
                "…" if len(vuln.summary) > 100 else ""
            )
            table_data.append(
                [
                    Paragraph(escape_html(vuln.vuln_id), st["table_cell"]),
                    Paragraph(escape_html(vuln.package_name), st["table_cell"]),
                    Paragraph(
                        f'<font color="{sev_color}">{vuln.severity.value}</font>',
                        st["table_cell"],
                    ),
                    Paragraph(escape_html(vuln.affected_versions), st["table_cell"]),
                    Paragraph(escape_html(summary_short), st["table_cell"]),
                ]
            )

        if len(table_data) > 1:
            row_count = len(table_data)
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), PDF_TABLE_HEADER_BG),
                        ("TEXTCOLOR", (0, 0), (-1, 0), PDF_TABLE_HEADER_TEXT),
                        ("FONTNAME", (0, 0), (-1, 0), FONT_RALEWAY_BOLD),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 0), (-1, 0), 8),
                        ("FONTNAME", (0, 1), (-1, -1), FONT_RALEWAY),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, row_count - 1),
                            [PDF_TABLE_ROW_ODD, PDF_TABLE_ROW_EVEN],
                        ),
                        ("GRID", (0, 0), (-1, -1), 0.3, PDF_BG_PAGE),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 1), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                        ("TEXTCOLOR", (0, 1), (-1, -1), PDF_TEXT_PRIMARY),
                    ]
                )
            )
            story.append(tbl)
        else:
            story.append(Paragraph("Keine Schwachstellen gefunden.", st["body_dim"]))

        # Advisories ohne Versionsabgleich — getrennt ausgewiesen,
        # nicht in der Severity-Tabelle.
        if result.unverified_vulnerabilities:
            story.append(Spacer(1, 0.5 * cm))
            story.append(
                Paragraph("Version unbekannt — Abgleich nicht möglich", st["h2"])
            )
            # Package-Name + Advisory-ID sind untrusted →
            # vor dem Paragraph-Rendern escapen.
            unverified_entries = [
                f"{escape_html(v.package_name)} ({escape_html(v.vuln_id)})"
                for v in result.unverified_vulnerabilities[:_PDF_MAX_LIST_ITEMS]
            ]
            unverified_str = ", ".join(unverified_entries)
            rest = len(result.unverified_vulnerabilities) - _PDF_MAX_LIST_ITEMS
            if rest > 0:
                unverified_str += f" … (+ {rest} weitere)"
            story.append(Paragraph(unverified_str, st["body_dim"]))

        # Unpinned
        if result.unpinned_dependencies:
            story.append(Spacer(1, 0.5 * cm))
            story.append(Paragraph("Nicht-gepinnte Dependencies", st["h2"]))
            # Namen stammen aus untrusted requirements-Dateien.
            unpinned_str = ", ".join(
                escape_html(d.name)
                for d in result.unpinned_dependencies[:_PDF_MAX_LIST_ITEMS]
            )
            if len(result.unpinned_dependencies) > _PDF_MAX_LIST_ITEMS:
                rest_unpinned = len(result.unpinned_dependencies) - _PDF_MAX_LIST_ITEMS
                unpinned_str += f" … (+ {rest_unpinned} weitere)"
            story.append(Paragraph(unpinned_str, st["body_dim"]))

        builder.add_footer_page()
        builder.build()
        log.info(
            "Dependency-Audit PDF-Export: %d Vulns → %s",
            len(result.vulnerabilities),
            path,
        )
        return True
