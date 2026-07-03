"""
network_exporter — Export-Service für Netzwerk-Scanner-Ergebnisse.

Unterstützt JSON, Excel (.xlsx) und PDF (FINLAI Dark Theme).

Schichtzugehörigkeit: application/ — kein GUI-Import.

Author: Patrick Riederich
Version: 1.0
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from core import theme
from core.exceptions import ConfigurationError
from core.export.base_exporter import BaseExporter
from core.logger import get_logger

if TYPE_CHECKING:
    from tools.network_scanner.domain.models import NetworkScanResult

log = get_logger(__name__)

# Excel-Farben (ARGB)
_XL_TEAL = "FF26A69A"
_XL_WHITE = "FFFFFFFF"
_XL_ROW_ODD = "FF252525"
_XL_ROW_EVEN = "FF1E1E1E"
_XL_TEXT = "FFC8CCD0"

# Risikoklassen-Farben für PDF (Severity-Signal-Palette aus theme)
_RISK_COLORS = {
    "kritisch": theme.SEVERITY_SIGNAL_CRITICAL,
    "hoch": theme.SEVERITY_SIGNAL_HIGH,
    "mittel": theme.SEVERITY_SIGNAL_MEDIUM,
    "niedrig": theme.SEVERITY_SIGNAL_LOW,
    "info": theme.SEVERITY_SIGNAL_INFO,
}


class NetworkExporter(BaseExporter):
    """Exportiert NetworkScanResult-Objekte in JSON, XLSX und PDF.

    Attributes:
        _report_title: Titel für PDF-Reports.
    """

    _report_title = "Netzwerk-Scanner Report"

    @property
    def default_filename_stem(self) -> str:
        """Basis-Dateiname.

        Returns:
            String "netzwerk_scan_export".
        """
        return "netzwerk_scan_export"

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def export_json(self, data: object, path: str) -> bool:
        """Exportiert NetworkScanResult als JSON.

        Args:
            data: NetworkScanResult-Instanz.
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.
        """
        result: NetworkScanResult = data  # type: ignore[assignment]
        export_dt = datetime.now(UTC).isoformat()

        hosts_data = []
        for host in result.hosts:
            ports_data = [
                {
                    "port": p.port,
                    "state": p.state.value,
                    "service": p.service,
                    "banner": p.banner,
                    "risk": p.risk.value,
                    "hinweis": p.hinweis,
                }
                for p in host.offene_ports
            ]
            hosts_data.append(
                {
                    "host": host.host,
                    "erreichbar": host.erreichbar,
                    "betriebssystem": host.betriebssystem,
                    "scan_dauer_s": host.scan_dauer_s,
                    "offene_ports": ports_data,
                }
            )

        payload = {
            "meta": {
                "exported_at": export_dt,
                "generator": "NoRisk by FINLAI — Netzwerk-Scanner Export",
                "scan_id": result.scan_id,
                "ziel": result.ziel,
                "scanner_typ": result.scanner_typ,
                "gestartet_am": result.gestartet_am.isoformat(),
                "beendet_am": result.beendet_am.isoformat(),
                "dauer_s": result.dauer_s,
            },
            "zusammenfassung": {
                "gesamt_hosts": len(result.hosts),
                "erreichbare_hosts": len(result.erreichbare_hosts),
                "offene_ports_gesamt": result.anzahl_offene_ports,
            },
            "hosts": hosts_data,
        }
        Path(path).write_text(
            json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        log.info("Netzwerk-Scanner JSON-Export: %s Hosts → %s", len(result.hosts), path)
        return True

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    def export_xlsx(self, data: object, path: str) -> bool:
        """Exportiert NetworkScanResult als Excel-Datei.

        Zwei Sheets: "Offene Ports" (flache Tabelle) + "Host-Übersicht".

        Args:
            data: NetworkScanResult-Instanz.
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.

        Raises:
            RuntimeError: Bei fehlendem openpyxl.
        """
        try:
            import openpyxl  # noqa: PLC0415
            from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415
            from openpyxl.utils import get_column_letter  # noqa: PLC0415
        except ImportError as exc:
            raise ConfigurationError("openpyxl nicht installiert") from exc

        result: NetworkScanResult = data  # type: ignore[assignment]
        wb = openpyxl.Workbook()

        header_fill = PatternFill("solid", fgColor=_XL_TEAL)
        header_font = Font(name="Calibri", bold=True, color=_XL_WHITE)
        odd_fill = PatternFill("solid", fgColor=_XL_ROW_ODD)
        even_fill = PatternFill("solid", fgColor=_XL_ROW_EVEN)
        text_font = Font(name="Calibri", color=_XL_TEXT)

        # Sheet 1: Offene Ports (flach)
        ws1 = wb.active
        ws1.title = "Offene Ports"  # type: ignore[union-attr]
        headers1 = ["Host", "Port", "Dienst", "Risikoklasse", "Banner", "Hinweis"]
        for col_idx, h in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)  # type: ignore[union-attr]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 2
        for host in result.hosts:
            if not host.erreichbar:
                continue
            for port in host.offene_ports:
                fill = odd_fill if row_idx % 2 == 1 else even_fill
                values = [
                    host.host,
                    port.port,
                    port.service or "—",
                    port.risk.value,
                    port.banner[:80] if port.banner else "—",
                    port.hinweis[:120] if port.hinweis else "—",
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=val)  # type: ignore[union-attr]
                    cell.fill = fill
                    cell.font = text_font
                row_idx += 1

        for col_idx, width in enumerate([22, 8, 15, 12, 50, 60], 1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]
        ws1.freeze_panes = "A2"  # type: ignore[union-attr]

        # Sheet 2: Host-Übersicht
        ws2 = wb.create_sheet("Host-Übersicht")
        headers2 = [
            "Host",
            "Erreichbar",
            "Offene Ports",
            "Max. Risiko",
            "OS",
            "Dauer (s)",
        ]
        for col_idx, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, host in enumerate(result.hosts, 2):
            fill = odd_fill if row_idx % 2 == 1 else even_fill
            values = [
                host.host,
                "Ja" if host.erreichbar else "Nein",
                len(host.offene_ports),
                host.max_risiko.value if host.erreichbar else "—",
                host.betriebssystem or "—",
                f"{host.scan_dauer_s:.1f}",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.font = text_font

        for col_idx, width in enumerate([22, 12, 14, 14, 25, 12], 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = width
        ws2.freeze_panes = "A2"

        wb.save(path)
        log.info("Netzwerk-Scanner XLSX-Export: %s", path)
        return True

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    def export_pdf(
        self,
        data: object,
        path: str,
        title: str = "",
        subtitle: str = "",
    ) -> bool:
        """Exportiert NetworkScanResult als Dark-Theme PDF.

        Args:
            data: NetworkScanResult-Instanz.
            path: Zieldateipfad.
            title: Optionaler Titel.
            subtitle: Optionaler Untertitel.

        Returns:
            True bei Erfolg.
        """
        from reportlab.lib.units import cm  # noqa: PLC0415
        from reportlab.platypus import (  # noqa: PLC0415
            Paragraph,
            Spacer,
            Table,
            TableStyle,
        )

        from core.pdf.pdf_colors import (  # noqa: PLC0415
            PDF_BG_PAGE,
            PDF_TABLE_HEADER_BG,
            PDF_TABLE_HEADER_TEXT,
            PDF_TABLE_ROW_EVEN,
            PDF_TABLE_ROW_ODD,
            PDF_TEXT_PRIMARY,
        )
        from core.pdf.pdf_fonts import FONT_RALEWAY, FONT_RALEWAY_BOLD  # noqa: PLC0415
        from core.pdf.pdf_report_builder import DarkReportBuilder  # noqa: PLC0415

        result: NetworkScanResult = data  # type: ignore[assignment]
        export_dt = datetime.now(UTC).strftime("%d.%m.%Y %H:%M UTC")

        _title = title or self._report_title
        _subtitle = subtitle or f"NoRisk by FINLAI  ·  {export_dt}"

        builder = DarkReportBuilder(output_path=path, title=_title, subtitle=_subtitle)
        builder.add_cover(date_str=export_dt)

        st = builder._styles  # noqa: SLF001
        story = builder._story  # noqa: SLF001

        # Zusammenfassung
        story.append(Paragraph("Scan-Zusammenfassung", st["h2"]))
        story.append(
            Paragraph(
                f"Ziel: {result.ziel}  |  Scanner: {result.scanner_typ}  |  "
                f"Hosts gescannt: {len(result.hosts)}  |  "
                f"Erreichbar: {len(result.erreichbare_hosts)}  |  "
                f"Offene Ports: {result.anzahl_offene_ports}",
                st["body_dim"],
            )
        )
        story.append(Spacer(1, 0.4 * cm))

        # Offene Ports Tabelle
        story.append(Paragraph("Offene Ports", st["h2"]))

        col_widths = [4.5 * cm, 2 * cm, 3 * cm, 3 * cm, 5 * cm]
        header_row = ["Host", "Port", "Dienst", "Risiko", "Hinweis"]
        table_data: list = [[Paragraph(h, st["table_header"]) for h in header_row]]

        for host in result.hosts:
            if not host.erreichbar:
                continue
            for port in host.offene_ports:
                risk_color_hex = _RISK_COLORS.get(
                    port.risk.value, theme.SEVERITY_SIGNAL_INFO
                )
                table_data.append(
                    [
                        Paragraph(host.host, st["table_cell"]),
                        Paragraph(str(port.port), st["table_cell_center"]),
                        Paragraph(port.service or "—", st["table_cell"]),
                        Paragraph(
                            f'<font color="{risk_color_hex}">{port.risk.value}</font>',
                            st["table_cell"],
                        ),
                        Paragraph(
                            port.hinweis[:80] if port.hinweis else "—", st["table_cell"]
                        ),
                    ]
                )

        if len(table_data) > 1:
            row_count = len(table_data)
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), PDF_TABLE_HEADER_BG),
                        ("TEXTCOLOR", (0, 0), (-1, 0), PDF_TABLE_HEADER_TEXT),
                        ("FONTNAME", (0, 0), (-1, 0), FONT_RALEWAY_BOLD),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("TOPPADDING", (0, 0), (-1, 0), 8),
                        ("FONTNAME", (0, 1), (-1, -1), FONT_RALEWAY),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, row_count - 1),
                            [PDF_TABLE_ROW_ODD, PDF_TABLE_ROW_EVEN],
                        ),
                        ("GRID", (0, 0), (-1, -1), 0.3, PDF_BG_PAGE),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 1), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                        ("TEXTCOLOR", (0, 1), (-1, -1), PDF_TEXT_PRIMARY),
                    ]
                )
            )
            story.append(tbl)
        else:
            story.append(Paragraph("Keine offenen Ports gefunden.", st["body_dim"]))

        builder.add_footer_page()
        builder.build()
        log.info("Netzwerk-Scanner PDF-Export: %s", path)
        return True
