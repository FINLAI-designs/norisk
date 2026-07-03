"""
system_exporter — Export-Service für System-Scanner-Ergebnisse.

Unterstützt JSON, Excel (.xlsx) und PDF (FINLAI Dark Theme).

Schichtzugehörigkeit: application/ — kein GUI-Import.

Author: Patrick Riederich
Version: 1.0
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from core import theme
from core.exceptions import ConfigurationError
from core.export.base_exporter import BaseExporter
from core.logger import get_logger

if TYPE_CHECKING:
    from tools.system_scanner.domain.entities import ScanResult

log = get_logger(__name__)

# Excel-Farben (ARGB)
_XL_TEAL = "FF26A69A"
_XL_WHITE = "FFFFFFFF"
_XL_ROW_ODD = "FF252525"
_XL_ROW_EVEN = "FF1E1E1E"
_XL_TEXT = "FFC8CCD0"

# Severvity → Risiko-String für PDF
_STATUS_LABEL = {
    "active": "Aktiv",
    "inactive": "Inaktiv",
    "outdated": "Veraltet",
    "risk": "Risiko",
    "unknown": "Unbekannt",
}


class SystemExporter(BaseExporter):
    """Exportiert ScanResult-Objekte in JSON, XLSX und PDF.

    Attributes:
        _report_title: Titel für PDF-Reports.
    """

    _report_title = "System-Scanner Report"

    @property
    def default_filename_stem(self) -> str:
        """Basis-Dateiname.

        Returns:
            String "system_scan_export".
        """
        return "system_scan_export"

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------

    def export_json(self, data: object, path: str) -> bool:
        """Exportiert ScanResult als JSON.

        Args:
            data: ScanResult-Instanz.
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.

        Raises:
            TypeError: Wenn data kein ScanResult ist.
        """
        result: ScanResult = data  # type: ignore[assignment]
        export_dt = datetime.now(UTC).isoformat()

        payload = {
            "meta": {
                "exported_at": export_dt,
                "generator": "NoRisk by FINLAI — System-Scanner Export",
                "scan_id": result.scan_id,
                "scan_timestamp": result.timestamp.isoformat(),
            },
            "os_info": result.os_info.to_dict(),
            "security_components": [c.to_dict() for c in result.security_components],
            "software_list": [s.to_dict() for s in result.software_list],
            "scan_duration_s": result.scan_duration_s,
            "warnings": result.warnings,
        }
        Path(path).write_text(
            json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        log.info("System-Scanner JSON-Export: %s", path)
        return True

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    def export_xlsx(self, data: object, path: str) -> bool:
        """Exportiert ScanResult als Excel-Datei.

        Zwei Sheets: "Sicherheitskomponenten" + "OS-Info".

        Args:
            data: ScanResult-Instanz.
            path: Zieldateipfad.

        Returns:
            True bei Erfolg.

        Raises:
            RuntimeError: Bei fehlendem openpyxl.
        """
        try:
            import openpyxl  # noqa: PLC0415
            from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415
            from openpyxl.utils import get_column_letter  # noqa: PLC0415
        except ImportError as exc:
            raise ConfigurationError("openpyxl nicht installiert") from exc

        result: ScanResult = data  # type: ignore[assignment]
        wb = openpyxl.Workbook()

        header_fill = PatternFill("solid", fgColor=_XL_TEAL)
        header_font = Font(name="Calibri", bold=True, color=_XL_WHITE)
        odd_fill = PatternFill("solid", fgColor=_XL_ROW_ODD)
        even_fill = PatternFill("solid", fgColor=_XL_ROW_EVEN)
        text_font = Font(name="Calibri", color=_XL_TEXT)

        # Sheet 1: Sicherheitskomponenten
        ws1 = wb.active
        ws1.title = "Sicherheitskomponenten"  # type: ignore[union-attr]
        headers1 = ["Name", "Typ", "Status", "Version", "Letztes Update", "Details"]
        for col_idx, h in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=h)  # type: ignore[union-attr]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, comp in enumerate(result.security_components, 2):
            fill = odd_fill if row_idx % 2 == 1 else even_fill
            values = [
                comp.name,
                comp.type.value,
                _STATUS_LABEL.get(comp.status.value, comp.status.value),
                comp.version or "—",
                comp.last_updated or "—",
                comp.detail or "—",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)  # type: ignore[union-attr]
                cell.fill = fill
                cell.font = text_font

        for col_idx, width in enumerate([30, 20, 15, 15, 18, 40], 1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]
        ws1.freeze_panes = "A2"  # type: ignore[union-attr]

        # Sheet 2: OS-Info
        ws2 = wb.create_sheet("OS-Info")
        os = result.os_info
        ws2_rows = [
            ("Betriebssystem", os.name),
            ("Platform", os.platform.value),
            ("Version", os.version),
            ("Build", os.build),
            ("Architektur", os.architecture),
            ("Letztes Update", os.last_update or "—"),
            (
                "Update-Status",
                _STATUS_LABEL.get(os.update_status.value, os.update_status.value),
            ),
        ]
        for i, (key, val) in enumerate(ws2_rows, 1):
            k_cell = ws2.cell(row=i, column=1, value=key)
            k_cell.font = Font(name="Calibri", bold=True, color=_XL_TEAL)
            v_cell = ws2.cell(row=i, column=2, value=val)
            v_cell.font = text_font
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 40

        wb.save(path)
        log.info("System-Scanner XLSX-Export: %s", path)
        return True

    # ------------------------------------------------------------------
    # PDF
    # ------------------------------------------------------------------

    def export_pdf(
        self,
        data: object,
        path: str,
        title: str = "",
        subtitle: str = "",
    ) -> bool:
        """Exportiert ScanResult als Dark-Theme PDF.

        Args:
            data: ScanResult-Instanz.
            path: Zieldateipfad.
            title: Optionaler Titel (überschreibt Default).
            subtitle: Optionaler Untertitel.

        Returns:
            True bei Erfolg.
        """
        from reportlab.lib.units import cm  # noqa: PLC0415
        from reportlab.platypus import (  # noqa: PLC0415
            Paragraph,
            Spacer,
            Table,
            TableStyle,
        )

        from core.pdf.pdf_colors import (  # noqa: PLC0415
            PDF_BG_PAGE,
            PDF_TABLE_HEADER_BG,
            PDF_TABLE_HEADER_TEXT,
            PDF_TABLE_ROW_EVEN,
            PDF_TABLE_ROW_ODD,
            PDF_TEXT_PRIMARY,
        )
        from core.pdf.pdf_fonts import FONT_RALEWAY, FONT_RALEWAY_BOLD  # noqa: PLC0415
        from core.pdf.pdf_report_builder import DarkReportBuilder  # noqa: PLC0415

        result: ScanResult = data  # type: ignore[assignment]
        export_dt = datetime.now(UTC).strftime("%d.%m.%Y %H:%M UTC")

        _title = title or self._report_title
        _subtitle = subtitle or f"NoRisk by FINLAI  ·  {export_dt}"

        builder = DarkReportBuilder(output_path=path, title=_title, subtitle=_subtitle)
        builder.add_cover(date_str=export_dt)

        st = builder._styles  # noqa: SLF001
        story = builder._story  # noqa: SLF001

        # OS-Übersicht
        story.append(Paragraph("System-Übersicht", st["h2"]))
        os = result.os_info
        story.append(
            Paragraph(
                f"Betriebssystem: {os.name}  |  Version: {os.version}"
                f"  |  Architektur: {os.architecture}",
                st["body"],
            )
        )
        story.append(Spacer(1, 0.3 * cm))

        # Sicherheitskomponenten-Tabelle
        story.append(Paragraph("Sicherheitskomponenten", st["h2"]))

        col_widths = [5 * cm, 4 * cm, 3 * cm, 3 * cm, 2.5 * cm]
        header_row = ["Name", "Typ", "Status", "Version", "Update"]
        table_data: list = [[Paragraph(h, st["table_header"]) for h in header_row]]

        _status_colors = {
            "active": theme.SEVERITY_SIGNAL_OK,
            "inactive": theme.SEVERITY_SIGNAL_CRITICAL,
            "outdated": theme.SEVERITY_SIGNAL_MEDIUM,
            "risk": theme.SEVERITY_SIGNAL_HIGH,
            "unknown": theme.SEVERITY_SIGNAL_INFO,
        }

        for comp in result.security_components:
            scolor = _status_colors.get(comp.status.value, theme.SEVERITY_SIGNAL_INFO)
            status_label = _STATUS_LABEL.get(comp.status.value, comp.status.value)
            table_data.append(
                [
                    Paragraph(comp.name, st["table_cell"]),
                    Paragraph(comp.type.value, st["table_cell"]),
                    Paragraph(
                        f'<font color="{scolor}">{status_label}</font>',
                        st["table_cell"],
                    ),
                    Paragraph(comp.version or "—", st["table_cell"]),
                    Paragraph(
                        comp.last_updated[:10] if comp.last_updated else "—",
                        st["table_cell"],
                    ),
                ]
            )

        if len(table_data) > 1:
            row_count = len(table_data)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), PDF_TABLE_HEADER_BG),
                ("TEXTCOLOR", (0, 0), (-1, 0), PDF_TABLE_HEADER_TEXT),
                ("FONTNAME", (0, 0), (-1, 0), FONT_RALEWAY_BOLD),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                ("FONTNAME", (0, 1), (-1, -1), FONT_RALEWAY),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, row_count - 1),
                    [PDF_TABLE_ROW_ODD, PDF_TABLE_ROW_EVEN],
                ),
                ("GRID", (0, 0), (-1, -1), 0.3, PDF_BG_PAGE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 1), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                ("TEXTCOLOR", (0, 1), (-1, -1), PDF_TEXT_PRIMARY),
            ]
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle(style_cmds))
            story.append(tbl)

        builder.add_footer_page()
        builder.build()
        log.info("System-Scanner PDF-Export: %s", path)
        return True
